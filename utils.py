import configparser
import csv
import datetime
import os
import sqlite3
import sys
from pathlib import Path
from tkinter import messagebox

from openpyxl import Workbook, load_workbook

import messages
import scorecalc
from logman import logger


#Form a query to search based on Roll number and optionally Name
def formSearchQuery(tablename, studentRoll, studentName):
    query = "SELECT * from " + tablename + " WHERE Roll_Number=" + studentRoll

    if(studentName != None and studentName != ""):
        query = query + " AND Name=" + studentName

    query = query + ";"
    return query


#Fetch the Student Record given the Roll number - Primary key
def getStudentRecord(dbconn, currtable, studentRoll, studentName):

    searchQuery = formSearchQuery(currtable, studentRoll, studentName)
    logger.info(messages.STUDENT_SEARCH_QUERY.format(searchQuery))
    
    try:
        searchcursor = dbconn.cursor()
        searchcursor.execute(searchQuery)
    
        studentRecords = searchcursor.fetchall()

        if(len(studentRecords) == 0):
            logger.warning(messages.RECORD_NOT_FOUND)
            messagebox.showerror(title="RECORD_NOT_FOUND", message=messages.RECORD_NOT_FOUND)
            return None

        if(len(studentRecords) > 1):
            messagebox.showerror(title="DUPLICATE_RECORDS", message=messages.DUPLICATE_RECORDS)
            return None

        studentRecord = studentRecords[0]
        
        logger.info(messages.OBTAINED_RECORD.format(str(studentRecord)))
        
        return studentRecord
    
    except sqlite3.Error as e:
        messagebox.showerror(title="SEARCH_FAILED", message=messages.SEARCH_FAILED)
        logger.critical(messages.SEARCH_FAILED.format(searchQuery))
        logger.critical(e, exc_info=True)
        return None
    except:
        messagebox.showerror(title="UNEXPECTED_ERROR", message=messages.UNEXPECTED_ERROR)
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        return None


# Connects with DB, throws an error if not possible
def connectwithDB(dbpath):
    try:
        logger.info((messages.DB_TRY_CONNECTION).format(dbpath))
        dbconn = sqlite3.connect(dbpath)
        dbconn.commit()
    except sqlite3.Error as e:
        logger.critical((messages.DB_CONNECTION_FAIL).format(dbpath))
        logger.critical(e, exc_info=True)
        messagebox.showerror(title="CONNECTION FAILED",
                             message=messages.DB_CONN_FAIL)
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        messagebox.showerror(title="CONNECTION FAILED",
                             message=messages.DB_CONN_FAIL)

    if dbconn:
        logger.info((messages.DB_CONNECTION_PASS).format(dbpath))
        return dbconn


# Creates a table in the DB if it does not exist
def createtable(dbconn, tablename, schemafile):

    logger.info(messages.CREATE_TABLE.format(tablename, schemafile))
    
    fielddict = {}
    for line in open(schemafile):
        if len(line) == 0 or line == '\n':
            continue
        if line[0] in ( '!', '#' ):
            continue
        field, datatype = line.split('=')
        fielddict[field.strip()] = datatype.strip()

    createquery = "CREATE TABLE IF NOT EXISTS " + tablename + "(" + os.linesep

    fields = list(fielddict.keys())
    datatypes = list(fielddict.values())
    numfields = len(fielddict.keys())
    for i in range(0, numfields-1):
        createquery = createquery + fields[i] + \
            " " + datatypes[i] + "," + os.linesep

    createquery = createquery + \
        fields[numfields-1] + " " + datatypes[numfields-1]
    createquery += (os.linesep + ");")
    
    logger.info(messages.CREATE_TABLE_QUERY.format(createquery))

    try:
        createcursor = dbconn.cursor()
        createcursor.execute(createquery)
        dbconn.commit()
    except sqlite3.Error as e:
        logger.critical(messages.CREATE_TABLE_FAILED.format(tablename))
        logger.critical(e, exc_info=True)
        return False
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        return False

    return True


# Get Schema from db and keep it ready
def getSchemaFromTable(conn, currtable):

    templist = []
    try:
        cur = conn.cursor()
        
        schemaPragma = "PRAGMA table_info('" + currtable + "')"
        logger.info(messages.SCHEMA_QUERY.format(schemaPragma))
        
        cur.execute(schemaPragma)
        schema = cur.fetchall()
        
        for row in schema:
            templist.append(row[1])
            
        st=" "
        logger.info(messages.FETCHED_SCHEMA.format(currtable, st.join(templist)))
      
    except sqlite3.Error as e:
        logger.critical(messages.SCHEMA_PRAGMA_FAILED.format(schemaPragma))
        logger.critical(e, exc_info=True)
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
    finally:
        return templist


# Called to export table to XLSX file
def sqltoexcel(dbconn, currtable, exportfilepath):

    schemaList = getSchemaFromTable(dbconn, currtable)
        
    logger.info(messages.EXPORT_STARTER.format(exportfilepath, schemaList))
        
    # from openpyxl import Workbook
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Appending Schema Row first
    ws.append(schemaList)

    try:
        cur = dbconn.cursor()
        cur.execute("SELECT * from " + currtable + ";")
        results = cur.fetchall()

        totalrows = len(results)
        exportedrows = 0
        
        for eachrow in results:
            ws.append(eachrow)
            exportedrows+=1

        # Save the file
        wb.save(exportfilepath)
        dbconn.commit()
        
    except sqlite3.Error as e:
        logger.critical(messages.DATA_EXPORT_FAILED.format(currtable, exportfilepath))
        logger.critical(e, exc_info=True)
        return False
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        return False

    return (totalrows == exportedrows),exportedrows


# Accepts the XLSX field names and returns the Column list for the SQLite table as mapped in fieldmap.properties
def getdbfields(excelfields, fieldmapfile):

    fieldmapdict = {}
    
    for line in open(fieldmapfile):
        if len(line) == 0 or line == '\n':
            continue
        if line[0] in ( '!', '#' ):
            continue
        tablefield, excelfield = line.split('=')
        fieldmapdict[excelfield.strip()] = tablefield.strip()
        
    dbfields = []
    irrelevantcolindices = []
    allkeys = fieldmapdict.keys()
    for i in range(0, len(excelfields)):
	    if excelfields[i] in allkeys:
		    dbfields.append(fieldmapdict.get(excelfields[i]))
	    else:
		    irrelevantcolindices.append(i)

    dbfields.append("Age")

    return dbfields, irrelevantcolindices


# Given the current tablename, it will fetch the Academic start date
def academicyearstartdate(currtable):

    tablelist = currtable.split('_')
    term = int(tablelist[3])
    year = int(tablelist[5])

    if term == 3:
        year -= 1

    startdate = datetime.date(year, 7, 1)
    return startdate


# Get the start of the academic year to calculate Age
def getage(currtable, dob):

    if dob == None:
        return None

    startdate = academicyearstartdate(currtable)

    if not isinstance(dob, datetime.datetime):
        dateofbirth = (datetime.datetime.strptime(dob, "%d-%b-%Y")).date()
    else:
        dateofbirth = dob.date()

    difference = startdate - dateofbirth
    return int(difference.days/365)


# Insert a row into the SQLite table
def rowinserter(rowtoinsert, dbconn, currtable, dbfieldlist, numrows):

    try:
        updatecursor = dbconn.cursor()
        
        insertquery = "INSERT INTO " + currtable + " " + str(tuple(dbfieldlist)) + " VALUES (" + "?, "*(len(dbfieldlist)-1) + "?" + ")"
        
        updatecursor.execute(insertquery, rowtoinsert)
        dbconn.commit()
    
    except sqlite3.IntegrityError as ie:
        logger.critical(messages.RECORD_EXISTS.format(currtable, rowtoinsert[dbfieldlist.index("Roll_Number")]))
        logger.critical(ie, exc_info=True)
        return numrows
    except sqlite3.Error as e:
        logger.critical(messages.INSERT_QUERY_FAILED.format(insertquery))
        logger.critical(e, exc_info=True)
        return numrows
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        return numrows
    
    return numrows + 1


# If the date read from Excel is not string and of Datetime type, this will make it a String
def processdate(dob):

    if isinstance(dob, datetime.datetime):
        logger.info(messages.DOBVALUE.format(dob,type(dob)))
        newformat = str(dob.strftime('%d-%b-%Y'))
        return newformat
    else:
        return dob


#Prepare the student record values for loading by checking for Integrity, presence of Primary Key etc
def preparevalues(datalist, irrelevantcolindices, dbfieldlist, currtable, rowstoload):
    
    dobindex = dbfieldlist.index("DOB")
    rollindex = dbfieldlist.index("Roll_Number")
    nameindex = dbfieldlist.index("Name")

    poppeditems = 0

    for eachindex in irrelevantcolindices:
        del datalist[eachindex-poppeditems]
        poppeditems = poppeditems + 1

    if datalist[rollindex] == None:
        logger.warning(messages.ROLL_NUM_EMPTY.format(rowstoload))
        return False

    if datalist[nameindex] == None:
        logger.warning(messages.NAME_EMPTY.format(rowstoload))
        return False

    if not datalist[dobindex] is None:
        datalist[dobindex] = processdate(datalist[dobindex])

    datarow = tuple(datalist)

    rowtoinsert = datarow + (getage(currtable, datarow[dobindex]),)
    return rowtoinsert


# This function readies the queries to load the data from Excel to SQL
def exceltosql(dbconn, currtable, loadfilename, fieldmapfile):

    logger.info(messages.START_LOADING.format(currtable, loadfilename, fieldmapfile))
    
    wb = load_workbook(filename=loadfilename, read_only=True)
    ws = wb.active

    colrowiter = ws.iter_rows(max_row=1, values_only=True)
    excelfields = next(colrowiter)

    dbfieldlist, irrelevantcolindices = getdbfields(excelfields, fieldmapfile)

    logger.info(messages.DB_FIELD_LIST.format(dbfieldlist))

    rowstoload = 0
    insertedrows = 0

    for datarow in ws.iter_rows(values_only=True, min_row=2):
        if not all(v is None for v in datarow):
            rowstoload+=1
            datalist = list(datarow)
            
            rowtoinsert = preparevalues(datalist, irrelevantcolindices, dbfieldlist, currtable, rowstoload)
           
            if rowtoinsert != False:
                insertedrows = rowinserter(rowtoinsert, dbconn, currtable, dbfieldlist, insertedrows)

    logger.info(messages.ROW_DIFF.format(rowstoload, insertedrows))
    return insertedrows, rowstoload


# Gets the sport event timings and returns query to update
def getUpdateQueries(height, fiftytime, eighthundredtime, shotputdist, longjumpdist, agilitytime, recordDict):

    updateQueries = []
    
    if(height != ""):
        userResponse = "no"
        if(recordDict["Height"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for Height as " +
                str(recordDict["Height"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Height"] == None or userResponse == "yes"):
            updateQueries.append("Height = " + str(height))
            recordDict["Height"] = float(height)
                
    if(fiftytime != ""):
        userResponse = "no"
        if(recordDict["Speed 50m Time"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for 50m Time as " +
                str(recordDict["Speed 50m Time"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Speed 50m Time"] == None or userResponse == "yes"):
            updateQueries.append("\"Speed 50m Time\" = " + str(fiftytime))
            score = scorecalc.fiftymetre_score(float(fiftytime))
            updateQueries.append("\"Speed Score\" = " + str(score))
            updateQueries.append("\"Speed Remarks\" = " + scorecalc.getRemarks(score))
            recordDict["Speed 50m Time"] = float(fiftytime)
            recordDict["Speed Score"] = int(score)
            
    if(eighthundredtime != ""):
        userResponse = "no"
        if(recordDict["Endurance 800m Time"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for 800m Time as " +
                str(recordDict["Endurance 800m Time"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Endurance 800m Time"] == None or userResponse == "yes"):
            updateQueries.append("\"Endurance 800m Time\" = " + str(eighthundredtime))
            score = scorecalc.eighthundredmetre_score(float(eighthundredtime))
            updateQueries.append("\"Endurance Score\" = " + str(score))
            updateQueries.append("\"Endurance Remarks\" = " + scorecalc.getRemarks(score))
            recordDict["Endurance 800m Time"] = float(eighthundredtime)
            recordDict["Endurance Score"] = int(score)
            
    if(shotputdist != ""):
        userResponse = "no"
        if(recordDict["Strength Shotput Distance"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for Shotput as " +
                str(recordDict["Strength Shotput Distance"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Strength Shotput Distance"] == None or userResponse == "yes"):
            updateQueries.append("\"Strength Shotput Distance\" = " + str(shotputdist))
            score = scorecalc.shotput_score(float(shotputdist))
            updateQueries.append("\"Strength Score\" = " + str(score))
            updateQueries.append("\"Strength Remarks\" = " + scorecalc.getRemarks(score))
            recordDict["Strength Shotput Distance"] = float(shotputdist)
            recordDict["Strength Score"] = int(score)

    if(longjumpdist != ""):
        userResponse = "no"
        if(recordDict["Explosive Longjump Distance"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for Long Jump as " +
                str(recordDict["Explosive Longjump Distance"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Explosive Longjump Distance"] == None or userResponse == "yes"):
            updateQueries.append(
                "\"Explosive Longjump Distance\" = " + str(longjumpdist))
            score = scorecalc.longjump_score(float(longjumpdist))
            updateQueries.append("\"Explosive Power Score\" = " + str(score))
            updateQueries.append("\"Explosive Power Remarks\" = " + scorecalc.getRemarks(score))
            recordDict["Explosive Longjump Distance"] = float(longjumpdist)
            recordDict["Explosive Power Score"] = int(score)

    if(agilitytime != ""):
        userResponse = "no"
        if(recordDict["Agility 60m Time"] != None):
            userResponse = messagebox.askquestion(title="VALUE EXISTS", message="The chosen student already has a value for Agility Time as " +
                str(recordDict["Agility 60m Time"]) + " Do you want to override that with the currently entered value?")
        if(recordDict["Agility 60m Time"] == None or userResponse == "yes"):
            updateQueries.append("\"Agility 60m Time\" = " + str(agilitytime))
            score = scorecalc.agilityScore(float(agilitytime))
            updateQueries.append("\"Agility Score\" = " + str(score))
            updateQueries.append("\"Agility Remarks\" = " + scorecalc.getRemarks(score))
            recordDict["Agility 60m Time"] = float(agilitytime)
            recordDict["Agility Score"] = int(score)
            
    return updateQueries, recordDict


#Forms the final update query based on all user entered values
def formUpdateQuery(tablename, updateQueries, recordDict):

	query = "UPDATE " + tablename + " "

	if updateQueries:
		query = query + "SET "

	for uq in updateQueries:
		query = query + uq + ", "
	
	if(query.endswith(", ")):
		query = query[0:len(query)-2]

	query = query + " " + "WHERE Roll_Number=" + str(recordDict["Roll_Number"])
	return query


#Gets the query and updates the student record
def updateTable(dbconn, updatequery):
    
    try:
        cur = dbconn.cursor()
        cur.execute(updatequery)
        dbconn.commit()
    except sqlite3.Error as e:
        logger.critical(messages.UPDATE_TABLE_FAILED.format(updatequery))
        logger.critical(e, exc_info=True)
        return False
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("Exception : ", exc_info=True)
        return False
        
    return True
    

#Updates the Total score of the student after updating individual scores
def updateTotalScore(tablename, recordDict):
    allsportscore = ["Speed Score", "Endurance Score", "Strength Score", "Explosive Power Score", "Agility Score"]

    total = 0
    for sportscore in allsportscore:
        if(recordDict[sportscore] != None and recordDict[sportscore] >=0 and recordDict[sportscore] <= 10):
            total = total + recordDict[sportscore]

    if total != 0:
        recordDict["Total Score"] = total
    
    totalscoreupdatequery = "UPDATE " + tablename + " SET \"Total Marks\" = " + str(total) + " WHERE Roll_Number=" + str(recordDict["Roll_Number"])

    return totalscoreupdatequery


# Destroys all the frames when switching pagess
def destroyframes(frames):
     # Delete all the frames created in the calling page
    for eachframe in frames:
        eachframe.destroy()


#Backup Progress Bar
def progress(status, remaining, total):
    print(f'Copied {total-remaining} of {total} pages...')


# Backs up the current DB just before quitting the App
def backup(prjroot, storagefolder, backupfolder):
    
    today = datetime.date.today()
    today = (str(today)).replace('-','_')
    
    onlydbs = [f for f in os.listdir(storagefolder) if os.path.isfile(os.path.join(storagefolder, f)) and f.endswith(".db")]
    
    currdatefolder = Path(backupfolder + os.path.sep + today)
    if not os.path.exists(currdatefolder):
        os.mkdir(currdatefolder)
    
    try:
        for eachdb in onlydbs:
            
            logger.info(messages.CURRENT_BACKUP_DB.format(eachdb))
            storagedbpath = Path(storagefolder + os.path.sep + eachdb)
            storagedbconn = sqlite3.connect(storagedbpath)
            
            backupdbpath = Path(backupfolder + os.path.sep + today + os.path.sep + eachdb)
            backupdbconn = sqlite3.connect(backupdbpath)
            logger.info(messages.BACKUP_TARGET.format(backupdbpath._str))

            storagedbconn.backup(backupdbconn, pages=0, progress=progress)
            
            storagedbconn.commit()
            backupdbconn.commit()
            
            storagedbconn.close()
            backupdbconn.close()
            
            logger.info(messages.BACKUP_SUCCESS.format(eachdb))
    except sqlite3.Error as e:
        logger.critical(messages.BACKUP_CREATION_FAILED.format(eachdb))
        logger.critical(e, exc_info=True)
        return False
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("Exception : ", exc_info=True)
        return False
    finally:
        if storagedbconn:
            storagedbconn.close()
        if backupdbconn:
            backupdbconn.close()
    

# Verify the Initial Setup & create the folder in case it was deleted
def verifySetup(prjroot, SERVER, CONFIG, BACKUPS, EXPORTS):
    
    try:
        serverpath = Path(prjroot + os.path.sep + SERVER)
        if not os.path.exists(serverpath):
            os.mkdir(serverpath)
            logger.info(messages.SERVER_FOLDER_CREATE)
        
        configpath = Path(prjroot + os.path.sep + CONFIG)
        if not os.path.exists(configpath):
            os.mkdir(configpath)
            logger.info(messages.CONFIG_FOLDER_CREATE)
            
        backuppath = Path(prjroot + os.path.sep + BACKUPS)
        if not os.path.exists(backuppath):
            os.mkdir(backuppath)
            logger.info(messages.BACKUPS_FOLDER_CREATE)
            
        exportpath = Path(prjroot + os.path.sep + EXPORTS)
        if not os.path.exists(exportpath):
            os.mkdir(exportpath)
            logger.info(messages.EXPORTS_FOLDER_CREATE)
    except:
        logger.error(messages.UNEXPECTED_ERROR)
        logger.error("StackTrace : ",exc_info=True)
        return False


#Fetch the path of the Logo bundled along with the Executable
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    logopath = os.path.join(base_path, relative_path) 
    logger.info("Logo "+str(logopath))

    return logopath